#imports, name them, track dependencies
import tkinter as tk
from tkinter import messagebox
from tkcalendar import DateEntry
from openpyxl import Workbook
from openpyxl import load_workbook
import os

base_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(base_dir, "Test.xlsx")
stored_names = None
due_dates = ["not defined", "not defined"]

#main class
class App(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        
        # Create a container frame where all other frames will be stacked
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        # Iterate through all the page classes you want to use
        for F in (MainPage, SettingsPage, EditPage):
            frame = F(container, self)
            self.frames[F] = frame
            # Place all frames in the same location (0, 0)
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(MainPage)

    def show_frame(self, cont):
        """Raises the selected frame to the top of the stacking order."""
        frame = self.frames[cont]
        frame.tkraise()
        self.geometry("1200x600")

#logic for main page
class MainPage(tk.Frame):
    def __init__(self,parent,controller):
        tk.Frame.__init__(self,parent)
        
        #Menu items
        menu_button = tk.Button(self, text="Open Settings", command=lambda:controller.show_frame(SettingsPage))
        menu_button.pack(side="top")
        
        edit_button = tk.Button(self, text="Edit Lists", command=lambda:controller.show_frame(EditPage))
        edit_button.pack(side="top")
        
        label = tk.Label(self, text="Main Page")
        label.pack(side="top")
        
        email_type_buttons = [1, 2, 3, 4]
        #Divs for centering bs
        main_container = tk.Frame(self)
        main_container.pack(expand=True)  # centers container in window

        left_frame = tk.Frame(main_container)
        left_frame.pack(side="left", padx=20)

        right_frame = tk.Frame(main_container)
        right_frame.pack(side="left", padx=20)
        
        # This variable now stores the selected radio value
        email_type_var = tk.StringVar(value=str(email_type_buttons[0]))
        
        # Label that displays current selection
        email_type_label = tk.Label(left_frame, text=f"Selected: {email_type_var.get()}")
        email_type_label.pack()
        
        # Function to update label automatically
        def update_selection():
            email_type_label.config(text=f"Selected: {email_type_var.get()}")
        
        r = 2
        for etype in email_type_buttons:
            tk.Radiobutton(
                left_frame,
                text=str(etype),
                variable=email_type_var,
                value=str(etype),
                command=update_selection
            ).pack()
            r += 1
        
        #Names input, specified by characters
        nlabel = tk.Label(left_frame, text="Put Names in here")
        nlabel.pack()
        
        nameBox = tk.Text(left_frame, height=5,width=30)
        nameBox.pack()
        
        commitNames = tk.Button(left_frame,text="Enter Names", command=lambda: get_names_from_list(nameBox))
        commitNames.pack()
        
        #Select many names
        selectAll = tk.Button(left_frame, text="Send to All", command=lambda: get_all_names())
        selectAll.pack()
        
        #Date selection
        mlabel = tk.Label(left_frame, text="Put due dates in here in order")
        mlabel.pack()
        
        dueDate1 = DateEntry(left_frame, width=10, background='darkblue', foreground='white', borderwidth=2, date_pattern='mm-dd-yyyy')
        dueDate1.pack()
        
        dueDate2 = DateEntry(left_frame, width=10, background='darkblue', foreground='white', borderwidth=2, date_pattern='mm-dd-yyyy')
        dueDate2.pack()
        
        setDates = tk.Button(left_frame, text="Set dates", command=lambda: set_dates(dueDate1,dueDate2))
        setDates.pack()
        
        #recipients box
        label = tk.Label(right_frame, text="Recipients")
        label.pack()
        
        emailsBox = tk.Text(right_frame, height=5, width=30)
        emailsBox.pack()
        
        #email generated box
        label = tk.Label(right_frame, text="Result Email")
        label.pack()
        
        messageSpace = tk.Text(right_frame, height=20, width=30)
        messageSpace.pack()
        
        #generate email button
        generate_button = tk.Button(right_frame,text="Generate Email", command= lambda: generate_email(email_type_var,messageSpace,emailsBox))
        generate_button.pack()
        

#logic that will help put names into the email
def generate_email(email_type_var, message_space, emails_box):
    #clear them out to be safe
    message_space.delete("1.0", tk.END)
    emails_box.delete("1.0",tk.END)
    global stored_names
    global due_dates
    if stored_names is None:
        count_issue()
        return -1
    names_list = list(stored_names.items())
    print(names_list)
    print(f"Selected email type: {email_type_var.get()} {type(email_type_var.get())}")
    generated_email=f""
    recipients = ""
    #This logic will split into 4 parts, based on what email template you want.
    #Deputy in constant
    match email_type_var.get():
        case "1":
            email_template = open("MONTHLYSHIFTSIGNUP")
            date=due_dates[0]
            date2=due_dates[1]
            month = date.month
            print(f"{date2}")
            for line in email_template:
                generated_email+=line
            generated_email.format(Month=month,Date=date, Date2=date2)
            email_template.close()
        case "2":
            email_template = open("WEEKLYSHIFTREMINDER")
            date="define later"
            for line in email_template:
                generated_email+=line
            email_template.close()
        case "3":
            email_template = open("WEEKLYSHIFTREMINDERINPERSON")
            date="define later"
            for line in email_template:
                generated_email+=line
            email_template.close()
        case "4":
            email_template = open("TEMPREMINDER")
            for line in email_template:
                generated_email+=line
            email_template.close()
            kv = names_list[0]
            key,value = kv
            Name = key
            generated_email = generated_email.format(Name=Name)
    #setups recipients
    for name,email in names_list:
        recipients += f"{email},"
        recipients.format(email=email)
    message_space.insert('1.0',generated_email)
    emails_box.insert('1.0',recipients)

def get_names_from_list(nameBox):
    input_val = nameBox.get("1.0", 'end-1c').lower()
    words = input_val.split()
    if len(words) % 2 != 0 or len(words)==0:
        name_error_popup()
        return -1
    names = [" ".join(words[i:i+2]) for i in range(0, len(words), 2)]
    name_to_email={}
    workbook = load_workbook(filename=file_path) #Change later to an actual document, up at the top
    ws = workbook.active
    # Iterate through all rows in the worksheet
    for row in ws.iter_rows(min_row=1):
        name_cell = row[0].value
        email_cell = row[1].value
        if name_cell and name_cell.lower() in names:
            name_to_email[name_cell] = email_cell#should access the email, may need to change
    
    print(name_to_email)
    missing_names = [n for n in names if n not in
                 [k.lower() for k in name_to_email.keys()]]

    if missing_names:
        length_issue(names)
    set_names(name_to_email)
    return name_to_email

def get_all_names():
    name_to_email={}
    workbook = load_workbook(filename=file_path) #Change later to an actual document, up at the top
    ws = workbook.active
    skipCount = 2
    iter=0
    # Iterate through all rows in the worksheet
    for row in ws.iter_rows(min_row=1):
        if iter<skipCount:
            iter+=1
            continue
        name_cell = row[0].value
        email_cell = row[1].value
        name_to_email[name_cell] = email_cell#should access the email, may need to change
    set_names(name_to_email)
    return name_to_email
    

def set_names(names):
    global stored_names
    stored_names = names

def set_dates(dueDate1,dueDate2):
    global due_dates
    due_dates[0] = dueDate1.get_date()
    due_dates[1] = dueDate2.get_date()
    

#end main page logic
#logic for names page
class SettingsPage(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)

        label = tk.Label(self, text="Settings Page")
        label.pack()

        button = tk.Button(
            self,
            text="Back to Main",
            command=lambda: controller.show_frame(MainPage)
        )
        button.pack()

class EditPage(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)

        label = tk.Label(self, text="Edit Page")
        label.pack()
        
        button = tk.Button(
            self,
            text="Back to Main",
            command=lambda: controller.show_frame(MainPage)
        )
        button.pack()

#
def name_error_popup():
    # Displays an information message box with an "OK" button
    messagebox.showinfo("Name Error", "Uneven number of words — cannot form full name pairs, Please enter names as FirstName LastName.")
    
def length_issue(names):
    #displays a information message saying that X name did not get entered to the names
    messagebox.showinfo("Entering error",f"Error, these names did not get registered{names}")

def count_issue():
    #displays a message box that says you need to enter names
    messagebox.showinfo("Recipient Error", "Please enter a set of recipiants to recieve this email.")
#sets up basics
#root = tk.Tk()
#Sets intitial window size to be greater than 0
#root.title("CFC Email Generator")
#root.geometry("400x300")
#root.minsize(200,200)

app = App()
app.mainloop()

# start the app
#root.mainloop()


